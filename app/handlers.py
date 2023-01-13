from dataclasses import dataclass
from operator import attrgetter
from pathlib import Path
from typing import Any, Union

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table

from app.services.imdb import Movie


@dataclass
class Handler:
    from_name: str
    to_name: str
    from_movies: set[Movie]
    to_movies: set[Movie]
    both: dict[str, dict]
    only_from: list[Movie]
    only_to: list[Movie]

    @property
    def from_count(self) -> int:
        return len(self.from_movies)

    @property
    def to_count(self) -> int:
        return len(self.to_movies)

    @property
    def both_count(self) -> int:
        return len(self.both)

    @property
    def only_from_count(self) -> int:
        return len(self.only_from)

    @property
    def only_to_count(self) -> int:
        return len(self.only_to)

    def handle(self) -> None:
        raise NotImplementedError


class CliHandler(Handler):
    def handle(self):
        table = Table()
        table.add_column("Statistic", style="cyan")
        table.add_column(self.from_name, style="green")
        table.add_column(self.to_name, style="green")

        table.add_row("Number of ratings", str(self.from_count), str(self.to_count))
        table.add_row("Both rated", str(self.both_count), str(self.both_count))
        table.add_row("Only rated", str(self.only_from_count), str(self.only_to_count))

        console = Console()
        console.print(table)


@dataclass
class SheetHandler(Handler):
    path: Union[Path, str]

    def __post_init__(self):
        if not isinstance(self.path, Path):
            self.path = Path(self.path)

    def handle(self):
        wb = Workbook()
        both_sheet = wb.active
        both_sheet.title = "Both rated"
        from_sheet = wb.create_sheet(f"{self.from_name} only")
        to_sheet = wb.create_sheet(f"{self.to_name} only")

        columns = (
            ("ID", 15),
            ("Title", 50),
            ("Year", 12),
            ("Genre", 30),
            ("Runtime", 15),
            (self.from_name, 15),
            (self.to_name, 15),
            ("Difference", 15),
        )
        movies = sorted(
            self.both.values(), key=attrgetter("rating_difference"), reverse=True
        )
        attributes = (
            "id",
            "title",
            "year",
            "genre",
            "runtime",
            "rating",
            "compare_rating",
            "rating_difference",
        )
        self._add_movies_to_sheet(both_sheet, columns, movies, attributes)

        columns = [
            ("ID", 15),
            ("Title", 50),
            ("Year", 12),
            ("Genre", 30),
            ("Runtime", 15),
            ("Rating", 12),
        ]
        attributes = ("id", "title", "year", "genre", "runtime", "rating")
        self._add_movies_to_sheet(from_sheet, columns, self.only_from, attributes)
        self._add_movies_to_sheet(to_sheet, columns, self.only_to, attributes)

        wb.save(self.path)

    @staticmethod
    def _add_movies_to_sheet(
        sheet: Any,
        columns: tuple[tuple[str, int]],
        movies: list[Movie],
        attributes: list[str],
    ) -> None:
        for i, column in enumerate(columns):
            letter = get_column_letter(i + 1)
            cell = sheet[f"{letter}1"]
            cell.value = column[0]
            cell.font = Font(bold=True)
            sheet.column_dimensions[letter].width = column[1]

        for movie in movies:
            values = [getattr(movie, a) for a in attributes]
            sheet.append(values)

        sheet.freeze_panes = sheet["A2"]
